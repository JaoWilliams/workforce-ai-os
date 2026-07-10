"""
Cálculo del reporte de horas trabajadas (mód. demo #108) y generación de sus
exports en Excel/PDF. Empareja marcaciones entrada->salida consecutivas por
empleado (heurística simple, sin manejo de turnos que cruzan medianoche —
mismo nivel de simplicidad que el resto del MVP).

El PDF reutiliza el mismo esquema de marca (logos TechSupport + tenant) que
el contrato de trabajo (ver core/contracts_pdf.py) para mantener consistencia
visual entre todos los documentos generados por el sistema.
"""
import io
import os
from datetime import date, datetime, time, timezone
from typing import Optional
from uuid import UUID

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from PIL import Image as PILImage
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.platypus import Flowable, HRFlowable, Image as RLImage, PageBreak, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
from sqlalchemy import select

from app.db.models import AttendanceRecord, Branch, Employee

LOGOS_DIR = os.path.join(os.path.dirname(os.path.dirname(__file__)), "assets", "logos")
TECHSUPPORT_LOGO = os.path.join(LOGOS_DIR, "TechSupport_clean.png")
BURGERKING_LOGO = os.path.join(LOGOS_DIR, "bk_logo_clean.png")

BROWN = colors.HexColor("#502314")
ORANGE = colors.HexColor("#FF8732")
CREAM = colors.HexColor("#F5F0E8")


async def compute_report_rows(session, start_date: date, end_date: date, branch_id: Optional[UUID] = None):
    start_dt = datetime.combine(start_date, time.min, tzinfo=timezone.utc)
    end_dt = datetime.combine(end_date, time.max, tzinfo=timezone.utc)

    query = (
        select(AttendanceRecord, Employee, Branch)
        .join(Employee, Employee.id == AttendanceRecord.employee_id)
        .join(Branch, Branch.id == Employee.branch_id)
        .where(AttendanceRecord.recorded_at >= start_dt, AttendanceRecord.recorded_at <= end_dt)
        .order_by(Employee.id, AttendanceRecord.recorded_at)
    )
    if branch_id is not None:
        query = query.where(Employee.branch_id == branch_id)

    result = await session.execute(query)
    rows = result.all()

    by_employee = {}
    for record, employee, branch in rows:
        key = employee.id
        if key not in by_employee:
            by_employee[key] = {
                "employee_id": employee.id,
                "employee_name": f"{employee.first_name} {employee.last_name}",
                "branch_id": branch.id,
                "branch_name": branch.name,
                "branch_accounting_account": branch.accounting_account,
                "records": [],
            }
        by_employee[key]["records"].append(record)

    report_rows = []
    for emp_data in by_employee.values():
        records = emp_data["records"]
        total_seconds = 0.0
        days = set()
        sessions = 0
        pending_entrada = None
        for r in records:
            if r.type == "entrada":
                pending_entrada = r
            elif r.type == "salida" and pending_entrada is not None:
                delta = (r.recorded_at - pending_entrada.recorded_at).total_seconds()
                if delta > 0:
                    total_seconds += delta
                    days.add(pending_entrada.recorded_at.date().isoformat())
                    sessions += 1
                pending_entrada = None

        report_rows.append({
            "employee_id": emp_data["employee_id"],
            "employee_name": emp_data["employee_name"],
            "branch_id": emp_data["branch_id"],
            "branch_name": emp_data["branch_name"],
            "branch_accounting_account": emp_data["branch_accounting_account"],
            "days_worked": len(days),
            "total_hours": round(total_seconds / 3600, 2),
            "total_sessions": sessions,
        })

    report_rows.sort(key=lambda x: x["employee_name"])
    return report_rows


def _scaled_image(path, target_height):
    img = PILImage.open(path)
    w, h = img.size
    ratio = w / h
    return RLImage(path, width=target_height * ratio, height=target_height)


def _build_header_logos(tenant_name: str):
    left_cell = _scaled_image(TECHSUPPORT_LOGO, 0.42 * inch)
    right_cell = ""
    if os.path.exists(BURGERKING_LOGO) and "burger" in tenant_name.lower():
        right_cell = _scaled_image(BURGERKING_LOGO, 0.52 * inch)
    header = Table([[left_cell, right_cell]], colWidths=[3.5 * inch, 2.5 * inch])
    header.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ALIGN", (0, 0), (0, 0), "LEFT"),
        ("ALIGN", (1, 0), (1, 0), "RIGHT"),
        ("LEFTPADDING", (0, 0), (-1, -1), 0),
        ("RIGHTPADDING", (0, 0), (-1, -1), 0),
        ("TOPPADDING", (0, 0), (-1, -1), 0),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
    ]))
    return header


class _Bookmark(Flowable):
    """Flowable invisible que crea un marcador de PDF (outline navegable) en la página actual."""

    def __init__(self, title, key):
        Flowable.__init__(self)
        self.title = title
        self.key = key

    def wrap(self, availWidth, availHeight):
        return (0, 0)

    def draw(self):
        self.canv.bookmarkPage(self.key)
        self.canv.addOutlineEntry(self.title, self.key, level=0, closed=True)


def build_report_xlsx(rows, start_date: date, end_date: date) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Horas trabajadas"

    ws.append([f"Reporte de horas trabajadas: {start_date.isoformat()} a {end_date.isoformat()}"])
    ws.merge_cells("A1:F1")
    ws["A1"].font = Font(bold=True, size=13)
    ws.append([])

    headers = ["Empleado", "Sucursal", "Cuenta contable", "Días trabajados", "Sesiones", "Horas totales"]
    ws.append(headers)
    header_row = ws.max_row
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=header_row, column=col)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="502314", end_color="502314", fill_type="solid")

    for r in rows:
        ws.append([
            r["employee_name"], r["branch_name"], r["branch_accounting_account"] or "—",
            r["days_worked"], r["total_sessions"], r["total_hours"],
        ])

    total_hours = sum(r["total_hours"] for r in rows)
    ws.append([])
    ws.append(["TOTAL", "", "", "", "", round(total_hours, 2)])
    ws.cell(row=ws.max_row, column=1).font = Font(bold=True)
    ws.cell(row=ws.max_row, column=6).font = Font(bold=True)

    widths = [28, 22, 16, 16, 12, 14]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_report_pdf(rows, start_date: date, end_date: date, tenant_name: str = "") -> bytes:
    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=letter,
        topMargin=0.6 * inch, bottomMargin=0.7 * inch,
        leftMargin=0.6 * inch, rightMargin=0.6 * inch,
    )
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("ReportTitle", parent=styles["Title"], fontSize=17, textColor=BROWN, spaceAfter=0)
    section_style = ParagraphStyle("Section", parent=styles["Heading2"], fontSize=13, textColor=BROWN, spaceBefore=0, spaceAfter=4)
    meta_style = ParagraphStyle("Meta", parent=styles["Normal"], fontSize=9.5, textColor=colors.HexColor("#666666"))
    small = ParagraphStyle("Small", parent=styles["Normal"], fontSize=7.5, textColor=colors.grey)
    hint_style = ParagraphStyle("Hint", parent=styles["Normal"], fontSize=8, textColor=colors.HexColor("#888888"), fontName="Helvetica-Oblique")
    stat_label = ParagraphStyle("StatLabel", parent=styles["Normal"], fontSize=8, textColor=colors.HexColor("#666666"))
    stat_value = ParagraphStyle("StatValue", parent=styles["Normal"], fontSize=16, textColor=BROWN, fontName="Helvetica-Bold")
    link_style = ParagraphStyle("Link", parent=styles["Normal"], fontSize=9, textColor=BROWN, fontName="Helvetica-Bold")

    total_hours = sum(r["total_hours"] for r in rows)
    total_days = sum(r["days_worked"] for r in rows)
    employee_count = len(rows)

    branches = {}
    for r in rows:
        b = branches.setdefault(r["branch_id"], {
            "branch_name": r["branch_name"],
            "branch_accounting_account": r["branch_accounting_account"],
            "employees": [],
        })
        b["employees"].append(r)
    branch_list = sorted(branches.items(), key=lambda kv: kv[1]["branch_name"])
    branch_count = len(branch_list)

    story = [
        _build_header_logos(tenant_name),
        Spacer(1, 0.22 * inch),
        _Bookmark("Resumen general", "top"),
        Paragraph("Reporte de Horas Trabajadas", title_style),
        HRFlowable(width="100%", thickness=2, color=ORANGE, spaceBefore=4, spaceAfter=8),
        Paragraph(f"<b>Empleador:</b> {tenant_name}", meta_style),
        Paragraph(f"<b>Período:</b> {start_date.isoformat()} a {end_date.isoformat()}", meta_style),
        Paragraph(f"<b>Generado:</b> {datetime.now().strftime('%Y-%m-%d %H:%M')}", meta_style),
        Spacer(1, 0.22 * inch),
    ]

    stats_data = [
        [Paragraph("EMPLEADOS", stat_label), Paragraph("SUCURSALES", stat_label),
         Paragraph("DÍAS TRABAJADOS", stat_label), Paragraph("HORAS TOTALES", stat_label)],
        [Paragraph(str(employee_count), stat_value), Paragraph(str(branch_count), stat_value),
         Paragraph(str(total_days), stat_value), Paragraph(f"{total_hours:.2f}", stat_value)],
    ]
    stats_table = Table(stats_data, colWidths=[1.6 * inch] * 4)
    stats_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), CREAM),
        ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#E5DCC8")),
        ("INNERGRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#E5DCC8")),
        ("TOPPADDING", (0, 0), (-1, 0), 8),
        ("BOTTOMPADDING", (0, 0), (-1, 0), 2),
        ("TOPPADDING", (0, 1), (-1, 1), 0),
        ("BOTTOMPADDING", (0, 1), (-1, 1), 8),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
    ]))
    story.append(stats_table)
    story.append(Spacer(1, 0.3 * inch))

    story.append(Paragraph("Resumen por Sucursal", section_style))
    story.append(Paragraph("Click en una sucursal para ver el detalle por empleado (o abra el panel de marcadores del PDF).", hint_style))
    story.append(Spacer(1, 0.12 * inch))

    summary_data = [["Sucursal", "Cuenta contable", "Empleados", "Días", "Horas"]]
    for branch_id, b in branch_list:
        key = f"branch_{branch_id}"
        branch_hours = sum(e["total_hours"] for e in b["employees"])
        branch_days = sum(e["days_worked"] for e in b["employees"])
        link_para = Paragraph(f'<a href="#{key}" color="#502314"><u>{b["branch_name"]}</u></a>', link_style)
        summary_data.append([
            link_para, b["branch_accounting_account"] or "—",
            str(len(b["employees"])), str(branch_days), f"{branch_hours:.2f}",
        ])
    summary_data.append(["TOTAL", "", str(employee_count), str(total_days), f"{total_hours:.2f}"])

    summary_table = Table(summary_data, colWidths=[1.9 * inch, 1.4 * inch, 1.0 * inch, 0.8 * inch, 0.9 * inch])
    summary_style = [
        ("BACKGROUND", (0, 0), (-1, 0), BROWN),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ALIGN", (2, 0), (-1, -1), "RIGHT"),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#DDDDDD")),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
        ("LINEABOVE", (0, -1), (-1, -1), 1, BROWN),
        ("BACKGROUND", (0, -1), (-1, -1), CREAM),
    ]
    for i in range(1, len(summary_data) - 1):
        if i % 2 == 0:
            summary_style.append(("BACKGROUND", (0, i), (-1, i), colors.HexColor("#FAF8F4")))
    summary_table.setStyle(TableStyle(summary_style))
    story.append(summary_table)
    story.append(PageBreak())

    story.append(Paragraph("Detalle por Empleado", section_style))
    story.append(HRFlowable(width="100%", thickness=1, color=colors.HexColor("#DDDDDD"), spaceBefore=2, spaceAfter=14))

    for branch_id, b in branch_list:
        key = f"branch_{branch_id}"
        acct = b["branch_accounting_account"] or "—"
        story.append(_Bookmark(f'Sucursal: {b["branch_name"]}', key))
        story.append(Paragraph(b["branch_name"], section_style))
        story.append(Paragraph(f"<b>Cuenta contable:</b> {acct}", meta_style))
        story.append(Spacer(1, 0.1 * inch))

        detail_data = [["Empleado", "Días", "Sesiones", "Horas"]]
        for e in sorted(b["employees"], key=lambda x: x["employee_name"]):
            detail_data.append([
                e["employee_name"], str(e["days_worked"]), str(e["total_sessions"]), f'{e["total_hours"]:.2f}',
            ])
        branch_hours = sum(e["total_hours"] for e in b["employees"])
        detail_data.append(["Subtotal sucursal", "", "", f"{branch_hours:.2f}"])

        detail_table = Table(detail_data, colWidths=[2.6 * inch, 0.9 * inch, 1.0 * inch, 1.0 * inch])
        detail_style = [
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#7A4A32")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 8.5),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("ALIGN", (1, 0), (-1, -1), "RIGHT"),
            ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#DDDDDD")),
            ("TOPPADDING", (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
            ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
            ("LINEABOVE", (0, -1), (-1, -1), 1, BROWN),
            ("BACKGROUND", (0, -1), (-1, -1), CREAM),
        ]
        for i in range(1, len(detail_data) - 1):
            if i % 2 == 0:
                detail_style.append(("BACKGROUND", (0, i), (-1, i), colors.HexColor("#FAF8F4")))
        detail_table.setStyle(TableStyle(detail_style))
        story.append(detail_table)
        story.append(Spacer(1, 0.3 * inch))

    story.append(HRFlowable(width="100%", thickness=0.5, color=colors.HexColor("#DDDDDD")))
    story.append(Spacer(1, 0.08 * inch))
    story.append(Paragraph(
        "Las horas se calculan emparejando marcaciones de entrada y salida consecutivas por empleado. "
        "Este reporte es una referencia operativa generada automáticamente; requiere validación antes de "
        "usarse como base de un asiento contable o de nómina oficial.",
        small,
    ))

    def _draw_footer(canvas_obj, doc_obj):
        canvas_obj.saveState()
        canvas_obj.setFont("Helvetica", 7.5)
        canvas_obj.setFillColor(colors.grey)
        canvas_obj.drawString(
            0.6 * inch, 0.4 * inch,
            f"WORKFORCE AI \u2014 Reporte de Horas Trabajadas \u2014 P\u00e1gina {doc_obj.page}",
        )
        if tenant_name:
            canvas_obj.drawRightString(letter[0] - 0.6 * inch, 0.4 * inch, tenant_name)
        canvas_obj.restoreState()

    doc.build(story, onFirstPage=_draw_footer, onLaterPages=_draw_footer)
    return buf.getvalue()
