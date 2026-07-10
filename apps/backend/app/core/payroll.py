"""
Cálculo de nómina bruta (horas trabajadas × tarifa por hora derivada del
contrato vigente del empleado). La tarifa por hora NO es un valor fijo ni
un supuesto legal de este sistema: se deriva dividiendo Contract.base_salary
entre las horas estándar configuradas por el TENANT para esa frecuencia de
pago (PayrollHoursConfig — módulo Catálogos), cargadas explícitamente por
el cliente/su contador según su jurisdicción y política interna real.

Si una frecuencia de pago no tiene horas configuradas todavía, este módulo
NO inventa un valor: marca esos contratos como "configuración pendiente"
(hours_config_missing=True) y no calcula bruto para ellos, hasta que se
cargue el número verificado en Catálogos → Configuración de Nómina.

Reutiliza compute_report_rows (horas trabajadas) y el mismo esquema de
marca/drilldown del PDF de reportes (core/attendance_report.py) para
mantener consistencia visual entre todos los documentos generados.
"""
import io
from datetime import date, datetime
from typing import Optional
from uuid import UUID

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.platypus import HRFlowable, PageBreak, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
from sqlalchemy import select

from app.core.attendance_report import BROWN, CREAM, ORANGE, _Bookmark, _build_header_logos, compute_report_rows
from app.core.holidays import compute_holiday_adjustments
from app.db.models import Contract, OvertimeApproval, PayrollConcept, PayrollHoursConfig


async def compute_payroll_rows(session, start_date: date, end_date: date, branch_id: Optional[UUID] = None):
    hours_rows = await compute_report_rows(session, start_date, end_date, branch_id)
    if not hours_rows:
        return []

    employee_ids = [r["employee_id"] for r in hours_rows]
    result = await session.execute(
        select(Contract).where(Contract.employee_id.in_(employee_ids)).order_by(Contract.start_date.desc())
    )
    contracts = result.scalars().all()
    latest_contract_by_employee = {}
    for c in contracts:
        if c.employee_id not in latest_contract_by_employee:
            latest_contract_by_employee[c.employee_id] = c

    config_result = await session.execute(select(PayrollHoursConfig))
    hours_by_frequency = {c.pay_frequency: float(c.standard_hours) for c in config_result.scalars().all()}

    overtime_result = await session.execute(
        select(OvertimeApproval).where(
            OvertimeApproval.employee_id.in_(employee_ids),
            OvertimeApproval.work_date >= start_date,
            OvertimeApproval.work_date <= end_date,
        )
    )
    overtime_by_employee = {}
    for ot in overtime_result.scalars().all():
        overtime_by_employee.setdefault(ot.employee_id, []).append(ot)

    concept_result = await session.execute(
        select(PayrollConcept).where(PayrollConcept.code == "HORAS_EXTRA", PayrollConcept.active.is_(True))
    )
    overtime_concept = concept_result.scalar_one_or_none()
    overtime_factor = float(overtime_concept.value) if overtime_concept else None

    holiday_adjustments = await compute_holiday_adjustments(session, employee_ids, start_date, end_date, branch_id)
    holiday_concept_result = await session.execute(
        select(PayrollConcept).where(PayrollConcept.code == "FERIADO_OBLIGATORIO_TRABAJADO", PayrollConcept.active.is_(True))
    )
    holiday_concept = holiday_concept_result.scalar_one_or_none()
    holiday_factor = float(holiday_concept.value) if holiday_concept else None

    payroll_rows = []
    for r in hours_rows:
        contract = latest_contract_by_employee.get(r["employee_id"])
        row = dict(r)
        if contract is None:
            row.update({
                "has_contract": False, "currency": None, "base_salary": None,
                "pay_frequency": None, "hourly_rate": None, "gross_pay": None,
                "hours_config_missing": False,
            })
        else:
            standard_hours = hours_by_frequency.get(contract.pay_frequency)
            if not standard_hours:
                row.update({
                    "has_contract": True, "currency": contract.currency,
                    "base_salary": float(contract.base_salary), "pay_frequency": contract.pay_frequency,
                    "hourly_rate": None, "gross_pay": None, "hours_config_missing": True,
                })
            else:
                hourly_rate = float(contract.base_salary) / standard_hours
                gross_pay = round(hourly_rate * row["total_hours"], 2)
                row.update({
                    "has_contract": True, "currency": contract.currency,
                    "base_salary": float(contract.base_salary), "pay_frequency": contract.pay_frequency,
                    "hourly_rate": round(hourly_rate, 4), "gross_pay": gross_pay,
                    "hours_config_missing": False,
                })
        ot_rows = overtime_by_employee.get(r["employee_id"], [])
        has_pending_overtime = any(o.status == "pending" for o in ot_rows)
        approved_extra_hours = round(sum(float(o.extra_hours) for o in ot_rows if o.status == "approved"), 2)
        row["overtime_extra_hours"] = approved_extra_hours
        row["overtime_pending"] = has_pending_overtime
        row["overtime_concept_missing"] = False
        row["overtime_surcharge"] = None
        if row["gross_pay"] is not None:
            if has_pending_overtime:
                row["gross_pay"] = None
            elif approved_extra_hours > 0:
                if overtime_factor is None:
                    row["gross_pay"] = None
                    row["overtime_concept_missing"] = True
                else:
                    surcharge = round(approved_extra_hours * row["hourly_rate"] * (overtime_factor - 1), 2)
                    row["overtime_surcharge"] = surcharge
                    row["gross_pay"] = round(row["gross_pay"] + surcharge, 2)
        holiday_entry = holiday_adjustments.get(r["employee_id"])
        row["holiday_unworked_pay"] = None
        row["holiday_worked_surcharge"] = None
        row["holiday_concept_missing"] = False
        if holiday_entry and row["gross_pay"] is not None:
            if holiday_entry["unworked_paid_hours"] > 0:
                unworked_pay = round(holiday_entry["unworked_paid_hours"] * row["hourly_rate"], 2)
                row["holiday_unworked_pay"] = unworked_pay
                row["gross_pay"] = round(row["gross_pay"] + unworked_pay, 2)
            if holiday_entry["worked_surcharge_hours"] > 0:
                if holiday_factor is None:
                    row["gross_pay"] = None
                    row["holiday_concept_missing"] = True
                else:
                    h_surcharge = round(holiday_entry["worked_surcharge_hours"] * row["hourly_rate"] * (holiday_factor - 1), 2)
                    row["holiday_worked_surcharge"] = h_surcharge
                    if row["gross_pay"] is not None:
                        row["gross_pay"] = round(row["gross_pay"] + h_surcharge, 2)
        payroll_rows.append(row)

    payroll_rows.sort(key=lambda x: x["employee_name"])
    return payroll_rows


def _currency_totals(rows):
    totals = {}
    for r in rows:
        if r["currency"]:
            totals[r["currency"]] = totals.get(r["currency"], 0.0) + (r["gross_pay"] or 0.0)
    return totals


def build_payroll_xlsx(rows, start_date: date, end_date: date) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Nomina bruta"

    ws.append([f"Nómina bruta: {start_date.isoformat()} a {end_date.isoformat()}"])
    ws.merge_cells("A1:I1")
    ws["A1"].font = Font(bold=True, size=13)
    ws.append([])

    headers = ["Empleado", "Sucursal", "Cuenta contable", "Frecuencia de pago",
               "Salario base", "Moneda", "Horas", "Tarifa/hora", "Bruto"]
    ws.append(headers)
    header_row = ws.max_row
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=header_row, column=col)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="502314", end_color="502314", fill_type="solid")

    for r in rows:
        if not r["has_contract"]:
            bruto_cell = "Sin contrato"
        elif r["hours_config_missing"]:
            bruto_cell = "Configuración pendiente"
        else:
            bruto_cell = r["gross_pay"]
        ws.append([
            r["employee_name"], r["branch_name"], r["branch_accounting_account"] or "—",
            r["pay_frequency"] or "Sin contrato",
            r["base_salary"] if r["base_salary"] is not None else "—",
            r["currency"] or "—", r["total_hours"],
            r["hourly_rate"] if r["hourly_rate"] is not None else "—",
            bruto_cell,
        ])

    ws.append([])
    for currency, total in _currency_totals(rows).items():
        ws.append([f"TOTAL {currency}", "", "", "", "", "", "", "", round(total, 2)])
        ws.cell(row=ws.max_row, column=1).font = Font(bold=True)
        ws.cell(row=ws.max_row, column=9).font = Font(bold=True)

    widths = [26, 20, 16, 16, 14, 8, 10, 12, 14]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_payroll_pdf(rows, start_date: date, end_date: date, tenant_name: str = "") -> bytes:
    from reportlab.lib.utils import ImageReader

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=letter,
        topMargin=0.6 * inch, bottomMargin=0.7 * inch,
        leftMargin=0.6 * inch, rightMargin=0.6 * inch,
    )
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("PayrollTitle", parent=styles["Title"], fontSize=17, textColor=BROWN, spaceAfter=0)
    section_style = ParagraphStyle("Section", parent=styles["Heading2"], fontSize=13, textColor=BROWN, spaceBefore=0, spaceAfter=4)
    meta_style = ParagraphStyle("Meta", parent=styles["Normal"], fontSize=9.5, textColor=colors.HexColor("#666666"))
    small = ParagraphStyle("Small", parent=styles["Normal"], fontSize=7.5, textColor=colors.grey)
    hint_style = ParagraphStyle("Hint", parent=styles["Normal"], fontSize=8, textColor=colors.HexColor("#888888"), fontName="Helvetica-Oblique")
    stat_label = ParagraphStyle("StatLabel", parent=styles["Normal"], fontSize=8, textColor=colors.HexColor("#666666"))
    stat_value = ParagraphStyle("StatValue", parent=styles["Normal"], fontSize=14, textColor=BROWN, fontName="Helvetica-Bold")
    link_style = ParagraphStyle("Link", parent=styles["Normal"], fontSize=9, textColor=BROWN, fontName="Helvetica-Bold")
    warn_style = ParagraphStyle("Warn", parent=styles["Normal"], fontSize=8, textColor=colors.HexColor("#B00020"), fontName="Helvetica-Bold")

    employee_count = len(rows)
    branch_count = len({r["branch_id"] for r in rows})
    total_hours = sum(r["total_hours"] for r in rows)
    currency_totals = _currency_totals(rows)
    missing_contract = sum(1 for r in rows if not r["has_contract"])
    missing_hours_config = sum(1 for r in rows if r["has_contract"] and r["hours_config_missing"])

    branches = {}
    for r in rows:
        b = branches.setdefault(r["branch_id"], {
            "branch_name": r["branch_name"],
            "branch_accounting_account": r["branch_accounting_account"],
            "employees": [],
        })
        b["employees"].append(r)
    branch_list = sorted(branches.items(), key=lambda kv: kv[1]["branch_name"])

    story = [
        _build_header_logos(tenant_name),
        Spacer(1, 0.22 * inch),
        _Bookmark("Resumen general", "top"),
        Paragraph("Nómina Bruta", title_style),
        HRFlowable(width="100%", thickness=2, color=ORANGE, spaceBefore=4, spaceAfter=8),
        Paragraph(f"<b>Empleador:</b> {tenant_name}", meta_style),
        Paragraph(f"<b>Período:</b> {start_date.isoformat()} a {end_date.isoformat()}", meta_style),
        Paragraph(f"<b>Generado:</b> {datetime.now().strftime('%Y-%m-%d %H:%M')}", meta_style),
        Spacer(1, 0.22 * inch),
    ]

    stat_cells_top = [Paragraph("EMPLEADOS", stat_label), Paragraph("SUCURSALES", stat_label), Paragraph("HORAS TOTALES", stat_label)]
    stat_cells_bottom = [Paragraph(str(employee_count), stat_value), Paragraph(str(branch_count), stat_value), Paragraph(f"{total_hours:.2f}", stat_value)]
    for currency, total in currency_totals.items():
        stat_cells_top.append(Paragraph(f"BRUTO {currency}", stat_label))
        stat_cells_bottom.append(Paragraph(f"{total:,.2f}", stat_value))

    col_w = 6.3 * inch / len(stat_cells_top)
    stats_table = Table([stat_cells_top, stat_cells_bottom], colWidths=[col_w] * len(stat_cells_top))
    stats_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), CREAM),
        ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#E5DCC8")),
        ("INNERGRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#E5DCC8")),
        ("TOPPADDING", (0, 0), (-1, 0), 8),
        ("BOTTOMPADDING", (0, 0), (-1, 0), 2),
        ("TOPPADDING", (0, 1), (-1, 1), 0),
        ("BOTTOMPADDING", (0, 1), (-1, 1), 8),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
    ]))
    story.append(stats_table)
    story.append(Spacer(1, 0.18 * inch))

    if missing_contract:
        story.append(Paragraph(
            f"⚠ {missing_contract} empleado(s) sin contrato registrado — no se pudo calcular su bruto (ver detalle).",
            warn_style,
        ))
        story.append(Spacer(1, 0.1 * inch))
    if missing_hours_config:
        story.append(Paragraph(
            f"⚠ {missing_hours_config} empleado(s) con frecuencia de pago sin horas configuradas — "
            f"configure las horas estándar en Catálogos → Configuración de Nómina para calcular su bruto.",
            warn_style,
        ))
        story.append(Spacer(1, 0.15 * inch))

    story.append(Paragraph("Resumen por Sucursal", section_style))
    story.append(Paragraph("Click en una sucursal para ver el detalle por empleado (o abra el panel de marcadores del PDF).", hint_style))
    story.append(Spacer(1, 0.12 * inch))

    summary_data = [["Sucursal", "Cuenta contable", "Empleados", "Horas", "Bruto"]]
    for branch_id, b in branch_list:
        key = f"branch_{branch_id}"
        branch_hours = sum(e["total_hours"] for e in b["employees"])
        branch_totals = _currency_totals(b["employees"])
        gross_text = "<br/>".join(f"{cur} {tot:,.2f}" for cur, tot in branch_totals.items()) or "—"
        link_para = Paragraph(f'<a href="#{key}" color="#502314"><u>{b["branch_name"]}</u></a>', link_style)
        summary_data.append([
            link_para, b["branch_accounting_account"] or "—",
            str(len(b["employees"])), f"{branch_hours:.2f}", Paragraph(gross_text, meta_style),
        ])
    total_row = ["TOTAL", "", str(employee_count), f"{total_hours:.2f}",
                 Paragraph("<br/>".join(f"{cur} {tot:,.2f}" for cur, tot in currency_totals.items()) or "—", meta_style)]
    summary_data.append(total_row)

    summary_table = Table(summary_data, colWidths=[1.7 * inch, 1.3 * inch, 0.9 * inch, 0.8 * inch, 1.5 * inch])
    summary_style = [
        ("BACKGROUND", (0, 0), (-1, 0), BROWN),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ALIGN", (2, 0), (-1, -1), "RIGHT"),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#DDDDDD")),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
        ("LINEABOVE", (0, -1), (-1, -1), 1, BROWN),
        ("BACKGROUND", (0, -1), (-1, -1), CREAM),
    ]
    for i in range(1, len(summary_data) - 1):
        if i % 2 == 0:
            summary_style.append(("BACKGROUND", (0, i), (-1, i), colors.HexColor("#FAF8F4")))
    summary_table.setStyle(TableStyle(summary_style))
    story.append(summary_table)
    story.append(PageBreak())

    story.append(Paragraph("Detalle por Empleado", section_style))
    story.append(HRFlowable(width="100%", thickness=1, color=colors.HexColor("#DDDDDD"), spaceBefore=2, spaceAfter=14))

    for branch_id, b in branch_list:
        key = f"branch_{branch_id}"
        acct = b["branch_accounting_account"] or "—"
        story.append(_Bookmark(f'Sucursal: {b["branch_name"]}', key))
        story.append(Paragraph(b["branch_name"], section_style))
        story.append(Paragraph(f"<b>Cuenta contable:</b> {acct}", meta_style))
        story.append(Spacer(1, 0.1 * inch))

        detail_data = [["Empleado", "Frecuencia", "Salario base", "Horas", "Tarifa/hora", "Bruto"]]
        for e in sorted(b["employees"], key=lambda x: x["employee_name"]):
            if not e["has_contract"]:
                salario = tarifa = "—"
                bruto = "Sin contrato"
                freq = "—"
            elif e["hours_config_missing"]:
                salario = f'{e["currency"]} {e["base_salary"]:,.2f}'
                tarifa = "—"
                bruto = "Configuración pendiente"
                freq = e["pay_frequency"]
            else:
                salario = f'{e["currency"]} {e["base_salary"]:,.2f}'
                tarifa = f'{e["hourly_rate"]:,.2f}'
                bruto = f'{e["currency"]} {e["gross_pay"]:,.2f}'
                freq = e["pay_frequency"]
            detail_data.append([e["employee_name"], freq, salario, f'{e["total_hours"]:.2f}', tarifa, bruto])

        detail_table = Table(detail_data, colWidths=[1.9 * inch, 0.9 * inch, 1.1 * inch, 0.7 * inch, 0.9 * inch, 1.0 * inch])
        detail_style = [
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#7A4A32")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("ALIGN", (2, 0), (-1, -1), "RIGHT"),
            ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#DDDDDD")),
            ("TOPPADDING", (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ]
        for i in range(1, len(detail_data)):
            if i % 2 == 0:
                detail_style.append(("BACKGROUND", (0, i), (-1, i), colors.HexColor("#FAF8F4")))
        detail_table.setStyle(TableStyle(detail_style))
        story.append(detail_table)
        story.append(Spacer(1, 0.3 * inch))

    story.append(HRFlowable(width="100%", thickness=0.5, color=colors.HexColor("#DDDDDD")))
    story.append(Spacer(1, 0.08 * inch))
    story.append(Paragraph(
        "Cálculo de referencia: tarifa por hora = salario del contrato vigente ÷ horas estándar "
        "configuradas por el cliente para esa frecuencia de pago (Catálogos → Configuración de Nómina). "
        "Contratos con frecuencias sin configurar se marcan como \u201cconfiguración pendiente\u201d y no "
        "calculan bruto. No incluye cargas sociales, deducciones, horas extra ni conceptos del catálogo de "
        "nómina (aguinaldo, etc.). Requiere validación contable antes de usarse como planilla oficial.",
        small,
    ))

    def _ts_bk_readers():
        from app.core.attendance_report import BURGERKING_LOGO, TECHSUPPORT_LOGO
        import os
        ts = ImageReader(TECHSUPPORT_LOGO) if os.path.exists(TECHSUPPORT_LOGO) else None
        bk = ImageReader(BURGERKING_LOGO) if (os.path.exists(BURGERKING_LOGO) and "burger" in tenant_name.lower()) else None
        return ts, bk

    _ts_reader, _bk_reader = _ts_bk_readers()

    def _draw_footer(canvas_obj, doc_obj):
        canvas_obj.saveState()
        canvas_obj.setFont("Helvetica", 7.5)
        canvas_obj.setFillColor(colors.grey)
        canvas_obj.drawString(0.6 * inch, 0.4 * inch, f"WORKFORCE AI — Nómina Bruta — Página {doc_obj.page}")
        if tenant_name:
            canvas_obj.drawRightString(letter[0] - 0.6 * inch, 0.4 * inch, tenant_name)
        canvas_obj.restoreState()

    def _draw_mini_header(canvas_obj, doc_obj):
        canvas_obj.saveState()
        page_w, page_h = letter
        logo_h = 0.24 * inch
        x = page_w - 0.6 * inch
        y = page_h - 0.5 * inch
        if _bk_reader is not None:
            iw, ih = _bk_reader.getSize()
            w = logo_h * iw / ih
            x -= w
            canvas_obj.drawImage(_bk_reader, x, y, width=w, height=logo_h, mask="auto")
            x -= 0.12 * inch
        if _ts_reader is not None:
            iw, ih = _ts_reader.getSize()
            w = logo_h * iw / ih
            x -= w
            canvas_obj.drawImage(_ts_reader, x, y, width=w, height=logo_h, mask="auto")
        canvas_obj.restoreState()
        _draw_footer(canvas_obj, doc_obj)

    doc.build(story, onFirstPage=_draw_footer, onLaterPages=_draw_mini_header)
    return buf.getvalue()
